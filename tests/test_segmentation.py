"""Invariants that must survive the move from regex segmentation to a model.

Written with unittest rather than pytest so it runs on a clean install with nothing
extra:  python -m unittest discover -s tests

These are not exhaustive. They cover the things that, if they broke, would break quietly
and be found by reconciling a spreadsheet weeks later -- which is the failure mode this
whole change exists to remove.
"""

from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
sys.path.insert(0, str(ROOT / "scripts"))

import zalo_capture as zc                                          # noqa: E402
from lavabo import extras, flags, rawpaste, segment                # noqa: E402
from lavabo.extract import base                                    # noqa: E402


class FakeCompleter:
    def __init__(self, payload=None, error=None, finish=""):
        self.payload, self.error, self.finish = payload, error, finish
        self.max_tokens = None

    def complete_json(self, system, user, schema, *, max_tokens=0):
        self.max_tokens = max_tokens
        if self.error:
            raise self.error
        return segment.Completion(self.payload, 100, 20, self.finish)


CHAT = """13/7 đơn 5 (Chị Hương)
1 tủ BC52
Tổng 5.800
Đã cọc 500k
ok chị
Thu thêm 2.900"""
LINES = CHAT.splitlines()

# The model answers in line numbers only: 1-4 is the order, 6 is a revision of it.
ORDER = {"header_line": 1, "end_line": 4, "day": 13, "month": 7, "order_number": 5,
         "customer": "Chị Hương", "updates": []}


class SegmentParsing(unittest.TestCase):
    def test_text_is_sliced_from_the_paste_not_taken_from_the_model(self):
        """The model cannot paraphrase an order, because it never sends one back."""
        order = segment.parse_response({"orders": [ORDER]}, LINES).orders[0]
        self.assertEqual(order.body, "13/7 đơn 5 (Chị Hương)\n1 tủ BC52\n"
                                     "Tổng 5.800\nĐã cọc 500k")
        self.assertEqual(order.header, "13/7 đơn 5 (Chị Hương)")

    def test_a_bad_entry_costs_one_order_not_the_paste(self):
        result = segment.parse_response({"orders": [ORDER, {"header_line": 1}]}, LINES)
        self.assertEqual(len(result.orders), 1)
        self.assertEqual(result.rejected, 1)

    def test_numeric_strings_are_accepted(self):
        raw = dict(ORDER, day="13", month="7", order_number="5")
        self.assertEqual(
            segment.parse_response({"orders": [raw]}, LINES).orders[0].key, (13, 7, 5))

    def test_a_line_range_past_the_end_is_clamped_not_dropped(self):
        raw = dict(ORDER, end_line=9999)
        result = segment.parse_response({"orders": [raw]}, LINES)
        self.assertEqual(len(result.orders), 1)
        self.assertTrue(result.orders[0].body.endswith("Thu thêm 2.900"))

    def test_a_miscounted_header_line_is_refused(self):
        """Pointing at line 2 while claiming 13/7 đơn 5 would file a real order's lines
        under a key they do not belong to -- worse than not finding it at all."""
        raw = dict(ORDER, header_line=2)
        result = segment.parse_response({"orders": [raw]}, LINES)
        self.assertEqual(result.orders, [])
        self.assertEqual(result.miscounted, 1)

    def test_an_update_is_sliced_by_its_own_range(self):
        raw = dict(ORDER, updates=[{"start_line": 6, "end_line": 6, "confidence": "low"}])
        update = segment.parse_response({"orders": [raw]}, LINES).orders[0].updates[0]
        self.assertEqual(update.text, "Thu thêm 2.900")
        self.assertEqual(update.confidence, "low")

    def test_confidence_is_normalised(self):
        raw = dict(ORDER, updates=[{"start_line": 6, "end_line": 6,
                                    "confidence": "wildly wrong"}])
        update = segment.parse_response({"orders": [raw]}, LINES).orders[0].updates[0]
        self.assertEqual(update.confidence, "high")

    def test_provider_failure_never_raises(self):
        result = segment.segment(FakeCompleter(error=RuntimeError("429 quota")), "x", 7, 2026)
        self.assertFalse(result.ok)
        self.assertIn("429", result.error)

    def test_unparseable_response_is_an_error_not_a_crash(self):
        result = segment.segment(FakeCompleter(payload="not json"), "x", 7, 2026)
        self.assertFalse(result.ok)

    def test_a_truncated_answer_says_so(self):
        result = segment.segment(
            FakeCompleter(payload="{trunc", finish="MAX_TOKENS"), CHAT, 7, 2026)
        self.assertTrue(result.truncated)
        self.assertIn("cut off", result.error)


class OutputBudget(unittest.TestCase):
    """The first live run returned 1 order out of 69, having been given field
    extraction's 4096-token budget for a whole month of orders."""

    def test_the_budget_grows_with_the_paste(self):
        small = segment.output_budget(20)
        month = segment.output_budget(900)
        self.assertGreater(month, small)
        self.assertGreaterEqual(small, segment.MIN_OUTPUT_TOKENS)

    def test_a_real_month_gets_room_with_margin(self):
        """A measured run: ~900 lines, 69 orders, 4,952 output tokens."""
        self.assertGreaterEqual(segment.output_budget(900), 4952 * 1.5)

    def test_the_call_is_given_the_sized_budget_not_the_config_one(self):
        completer = FakeCompleter(payload={"orders": []})
        segment.segment(completer, "\n".join(["line"] * 900), 7, 2026)
        self.assertGreaterEqual(completer.max_tokens, segment.MIN_OUTPUT_TOKENS)


class IncompleteAnswers(unittest.TestCase):
    """A model that returns a tenth of the orders has failed, not disagreed. The log has
    to say which, because the first live run read like an ordinary disagreement."""

    def test_a_short_answer_is_called_a_failure(self):
        blocks = zc.split_orders("\n".join(
            f"{d}/7 đơn 1 - K{d}\n1 tủ\nTổng 5.800" for d in range(1, 21)), target_month=7)
        ai = segment.parse_response({"orders": [ORDER]}, LINES)
        text = segment.summarise(ai, blocks, segment.compare(ai, blocks))
        self.assertIn("INCOMPLETE", text)
        self.assertIn("do NOT switch", text)

    def test_a_matching_answer_says_nothing_alarming(self):
        blocks = zc.split_orders(CHAT, target_month=7)
        ai = segment.parse_response({"orders": [ORDER]}, LINES)
        text = segment.summarise(ai, blocks, segment.compare(ai, blocks))
        self.assertNotIn("INCOMPLETE", text)

    def test_truncation_is_reported_as_interruption(self):
        ai = segment.parse_response({"orders": [ORDER]}, LINES)
        ai.finish_reason = "MAX_TOKENS"
        self.assertIn("CUT OFF", segment.summarise(ai, [], []))


class BlockConversion(unittest.TestCase):
    """The model's output must become exactly what the regexes produce, because dedup,
    the month filter and the save path all run on those objects unchanged."""

    def test_key_matches_the_regex_splitter(self):
        by_regex = zc.split_orders(CHAT, target_month=7)
        by_model = zc.blocks_from_segments(
            segment.parse_response({"orders": [ORDER]}, LINES), target_month=7)
        self.assertEqual([b.key for b in by_regex], [b.key for b in by_model])
        self.assertEqual(by_regex[0].customer, by_model[0].customer)

    def test_the_header_is_not_repeated_inside_the_body(self):
        block = zc.blocks_from_segments(
            segment.parse_response({"orders": [ORDER]}, LINES))[0]
        self.assertNotIn(block.header, block.lines)
        self.assertIn("1 tủ BC52", block.lines)


class Comparison(unittest.TestCase):
    def test_a_date_the_deterministic_pass_fixes_is_not_a_disagreement(self):
        """The model read "8/3" literally where the splitter swapped it to 3 August. Both
        end up at 3/8 once converted, so the log must not report the order as missed by
        one side and invented by the other -- only_regex is the line that stops a switch."""
        chat = ("3/8 đơn 5 - Phương Phan\n1 tủ\nTổng 5.800\n"
                "8/3 đơn 1 - hoài bùi\n1 lavabo\nTổng 2tr\n"
                "3/8 đơn 2 - Nguyễn Tài Lợi\n1 sen\nTổng 3tr")
        lines = chat.splitlines()
        blocks = zc.split_orders(chat, target_month=8)
        ai = segment.parse_response({"orders": [
            {"header_line": 1, "end_line": 3, "day": 3, "month": 8, "order_number": 5},
            {"header_line": 4, "end_line": 6, "day": 8, "month": 3, "order_number": 1},
            {"header_line": 7, "end_line": 9, "day": 3, "month": 8, "order_number": 2},
        ]}, lines)

        raw = segment.compare(ai, blocks)
        self.assertEqual(sorted(d.kind for d in raw), ["only_ai", "only_regex"])

        converted = segment.compare(ai, blocks, zc.blocks_from_segments(ai, target_month=8))
        self.assertEqual(converted, [])


    def test_orders_only_one_side_found_are_reported(self):
        blocks = zc.split_orders(CHAT, target_month=7)
        lines = LINES + ["16/7 đơn 2", "1 lavabo", "Tổng 2tr"]
        extra = {"header_line": 7, "end_line": 9, "day": 16, "month": 7,
                 "order_number": 2}
        found = segment.compare(
            segment.parse_response({"orders": [ORDER, extra]}, lines), blocks)
        self.assertEqual([d.kind for d in found], ["only_ai"])

    def test_a_failed_segmentation_reports_as_error(self):
        self.assertEqual(
            [d.kind for d in segment.compare(segment.SegmentResult(error="boom"), [])],
            ["error"])


class Sidecars(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.inbox = Path(self.tmp.name) / "zalo"
        self.inbox.mkdir(parents=True)

    def tearDown(self):
        self.tmp.cleanup()

    def test_re_pasting_does_not_stack_the_same_revision(self):
        for _ in range(3):
            extras.record(self.inbox, "o.txt", "update", "Thu thêm 2.900")
        self.assertEqual(len(extras.load(self.inbox)["o.txt"]), 1)

    def test_a_confident_repeat_upgrades_an_uncertain_revision(self):
        extras.record(self.inbox, "o.txt", "update", "Tổng 12.000", "low")
        extras.record(self.inbox, "o.txt", "update", "Tổng 12.000", "high")
        self.assertEqual(extras.load(self.inbox)["o.txt"][0]["confidence"], "high")

    def test_flags_are_added_once_and_can_be_cleared(self):
        self.assertTrue(flags.record(self.inbox, "o.txt", flags.NO_AI))
        self.assertFalse(flags.record(self.inbox, "o.txt", flags.NO_AI))
        self.assertTrue(flags.clear(self.inbox, "o.txt", flags.NO_AI))
        self.assertEqual(flags.load(self.inbox), {})

    def test_a_damaged_sidecar_reads_as_empty_rather_than_raising(self):
        for module in (extras, flags):
            module.sidecar_path(self.inbox).write_text("{ broken", encoding="utf-8")
            self.assertEqual(module.load(self.inbox), {})

    def test_the_paste_is_stored_byte_for_byte_and_only_once(self):
        text = "13/7 đơn 5 (Chị Hương)\n1 tủ BC52\nTổng 5.800\nĐã cọc 500k\n"
        first = rawpaste.store(self.inbox, text, month=7, year=2026)
        again = rawpaste.store(self.inbox, text, month=7, year=2026)
        self.assertEqual(first, again)
        self.assertEqual(len(rawpaste.load_index(self.inbox)), 1)
        self.assertEqual(first.read_text(encoding="utf-8"), text)

    def test_the_raw_store_is_not_inside_the_inbox(self):
        """Anything under the inbox is walked by the connector and extracted as a
        transcript, so a paste left there would be ingested as a fake conversation."""
        self.assertNotIn(self.inbox, rawpaste.store_dir(self.inbox).parents)


class SegmentationSetting(unittest.TestCase):
    """The setting is chosen in the UI but lives in a YAML file people also hand-edit."""

    def test_yaml_turns_bare_on_and_off_into_booleans(self):
        """The trap this guards. If it ever stops being true, the coercion is dead code
        rather than a silent bug waiting -- so assert the premise, not just the fix."""
        import yaml
        self.assertIs(yaml.safe_load("x: on")["x"], True)
        self.assertIs(yaml.safe_load("x: off")["x"], False)

    def test_a_hand_written_mode_survives_that(self):
        from lavabo.config import _segmentation_mode
        self.assertEqual(_segmentation_mode(True), "on")
        self.assertEqual(_segmentation_mode(False), "off")
        self.assertEqual(_segmentation_mode("shadow"), "shadow")
        self.assertEqual(_segmentation_mode(" ON "), "on")

    def test_an_unrecognised_value_does_not_read_as_deliberate(self):
        from lavabo.config import _segmentation_mode
        with self.assertLogs("lavabo.config", level="WARNING"):
            self.assertEqual(_segmentation_mode("maybe"), "off")

    def test_the_loader_and_the_settings_screen_agree(self):
        """They read the same file by different paths, and a screen showing "off" while
        capture runs the model would be worse than either being wrong alone."""
        import lavabo.settings as settings_module
        from lavabo.config import Config

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "config.yaml"
            path.write_text("extract:\n  provider: gemini\n  ai_segmentation: on\n",
                            encoding="utf-8")
            original = settings_module.CONFIG_PATH
            settings_module.CONFIG_PATH = path
            try:
                self.assertEqual(settings_module.read_settings()["segmentation"], "on")
            finally:
                settings_module.CONFIG_PATH = original
            self.assertEqual(Config.load(path).extract.ai_segmentation, "on")


class OneBodyOfKnowledge(unittest.TestCase):
    """Text and video are two ways in for the same shop. What an order looks like must be
    stated once: two copies means fixing every future discovery twice, and finding out
    later that one copy was missed."""

    def test_both_prompts_carry_the_same_shop_context(self):
        from lavabo.extract.prompt import ORDER_CONTEXT, build_segment_prompt, build_video_prompt

        shared = ORDER_CONTEXT.format(month=7, year=2026)
        text_system = build_segment_prompt("13/7 đơn 5", 7, 2026)[0]
        video_system = build_video_prompt(12, 7, 2026)[0]
        self.assertIn(shared, text_system)
        self.assertIn(shared, video_system)

    def test_the_hard_won_details_reach_the_video_path(self):
        """Each of these cost a real bug on the text path."""
        from lavabo.extract.prompt import build_video_prompt

        system = build_video_prompt(12, 7, 2026)[0]
        for detail in ("gương cộc",             # a mirror, not a deposit
                       "(Trần Thị Liên)",        # bracketed header form
                       "8th of March",           # day-first dates
                       "1tr8"):                  # money as written, never converted
            with self.subTest(detail):
                self.assertIn(detail, system)

    def test_only_the_answer_contract_differs(self):
        from lavabo.extract.prompt import build_segment_prompt, build_video_prompt

        text_system = build_segment_prompt("13/7 đơn 5", 7, 2026)[0]
        video_system = build_video_prompt(12, 7, 2026)[0]
        self.assertIn("ONLY in line numbers", text_system)
        self.assertNotIn("ONLY in line numbers", video_system)
        self.assertIn("TRANSCRIBE", video_system)
        self.assertNotIn("TRANSCRIBE", text_system)


class ReadingFromVideo(unittest.TestCase):
    def _order(self, frame, day, number, body, partial=False):
        return {"frame": frame, "header": body.splitlines()[0], "day": day, "month": 8,
                "order_number": number, "body": body, "partial": partial}

    def test_the_same_order_filmed_repeatedly_is_one_order(self):
        """Consecutive frames overlap by design, so the model is told to report an order
        every time it sees it. Reconciling that is arithmetic, not judgement."""
        body = "15/8 đơn 1 - Meloxicam\n1 tủ BC52\nTổng 29tr"
        result = segment.parse_video_response({"orders": [
            self._order(1, 15, 1, body), self._order(2, 15, 1, body),
            self._order(3, 15, 1, body)]})
        self.assertEqual(len(result.orders), 1)

    def test_an_order_clipped_by_a_frame_edge_loses_to_the_whole_one(self):
        clipped = "15/8 đơn 1 - Meloxicam\n1 tủ BC52"
        whole = "15/8 đơn 1 - Meloxicam\n1 tủ BC52\nTổng 29tr\nĐã cọc 500k"
        for order in ([self._order(1, 15, 1, clipped, partial=True),
                       self._order(2, 15, 1, whole)],
                      [self._order(1, 15, 1, whole),
                       self._order(2, 15, 1, clipped, partial=True)]):
            with self.subTest(first=order[0]["partial"]):
                result = segment.parse_video_response({"orders": order})
                self.assertEqual(len(result.orders), 1)
                self.assertEqual(result.orders[0].body, whole)

    def test_a_clipped_order_is_kept_when_nothing_better_exists(self):
        """Half an order is worth reviewing; it just must not beat the whole one."""
        clipped = "15/8 đơn 1 - Meloxicam\n1 tủ BC52"
        result = segment.parse_video_response(
            {"orders": [self._order(1, 15, 1, clipped, partial=True)]})
        self.assertEqual(len(result.orders), 1)

    def test_an_unusable_row_costs_one_order_not_the_recording(self):
        result = segment.parse_video_response({"orders": [
            self._order(1, 15, 1, "15/8 đơn 1\nTổng 2tr"), {"frame": 2}]})
        self.assertEqual(len(result.orders), 1)
        self.assertEqual(result.rejected, 1)

    def test_a_failed_call_never_raises(self):
        class Boom:
            def complete_json_images(self, *a, **k):
                raise RuntimeError("503 unavailable")
        result = segment.segment_video(Boom(), [b"x"], 8, 2026)
        self.assertFalse(result.ok)

    def test_no_frames_is_an_error_not_a_call(self):
        self.assertFalse(segment.segment_video(None, [], 8, 2026).ok)


class Retrying(unittest.TestCase):
    """A 503 matched none of the rate-limit patterns and was not retried at all, and the
    Anthropic path had no retry of any kind. Either way a momentary provider hiccup
    dropped a whole paste to the regex fallback, losing the revisions the model is for."""

    class Clock:
        """A fake sleep that advances a fake clock, so a deadline test proves something."""
        def __init__(self):
            self.t, self.waits = 0.0, []

        def sleep(self, seconds):
            self.waits.append(seconds)
            self.t += seconds

        def now(self):
            return self.t

    def test_provider_hiccups_are_retried(self):
        for message in ("503 UNAVAILABLE: The model is overloaded",
                        "529 overloaded_error", "500 INTERNAL", "502 Bad Gateway",
                        "504 Deadline Exceeded", "ReadTimeout: timed out",
                        "Connection reset by peer"):
            with self.subTest(message):
                self.assertTrue(base.is_transient(Exception(message)))

    def test_quota_errors_are_told_apart_from_hiccups(self):
        """They need different waits: a quota is a promise about the next minute, an
        overload usually clears in seconds."""
        self.assertTrue(base.is_rate_limit(Exception("429 RESOURCE_EXHAUSTED quota")))
        self.assertFalse(base.is_rate_limit(Exception("503 unavailable")))

    def test_a_bad_request_is_not_retried(self):
        attempts = []

        def bad():
            attempts.append(1)
            raise ValueError("400 INVALID_ARGUMENT: bad schema")

        with self.assertRaises(ValueError):
            base.retry_call(bad, sleep=lambda s: None)
        self.assertEqual(len(attempts), 1)

    def test_a_burst_of_503s_recovers(self):
        attempts = []

        def flaky():
            attempts.append(1)
            if len(attempts) < 3:
                raise RuntimeError("503 UNAVAILABLE: model is overloaded")
            return "answer"

        clock = self.Clock()
        self.assertEqual(
            base.retry_call(flaky, sleep=clock.sleep, monotonic=clock.now), "answer")
        self.assertEqual(len(attempts), 3)
        self.assertLess(clock.t, 30, "an overload should not cost half a minute")

    def test_an_overload_waits_far_less_than_a_quota(self):
        totals = {}
        for label, message in (("quota", "429 quota exceeded"),
                               ("overload", "503 unavailable")):
            clock = self.Clock()
            with self.assertRaises(RuntimeError):
                base.retry_call(
                    lambda: (_ for _ in ()).throw(RuntimeError(message)),
                    sleep=clock.sleep, monotonic=clock.now)
            totals[label] = clock.t
        self.assertGreater(totals["quota"], totals["overload"] * 5)

    def test_the_deadline_is_respected_not_merely_checked(self):
        """Segmentation runs while somebody waits for Lưu đơn, so it must not spin for
        the five minutes extraction is allowed."""
        clock = self.Clock()
        with self.assertRaises(RuntimeError):
            base.retry_call(
                lambda: (_ for _ in ()).throw(RuntimeError("429 quota exceeded")),
                deadline_seconds=90, sleep=clock.sleep, monotonic=clock.now)
        self.assertLessEqual(clock.t, 90)

    def test_extraction_keeps_trying_for_minutes(self):
        """No deadline: nobody is watching a batch, and a retry is cheaper than a blank
        column somebody has to notice."""
        clock = self.Clock()
        with self.assertRaises(RuntimeError):
            base.retry_call(
                lambda: (_ for _ in ()).throw(RuntimeError("429 quota exceeded")),
                sleep=clock.sleep, monotonic=clock.now)
        self.assertGreater(clock.t, 240)


class MoneyNeverMoves(unittest.TestCase):
    """The one invariant worth more than all the others: a revision is shown, never
    applied. The shop reconciles totals by hand against its own workbook."""

    def test_a_revision_stays_out_of_the_total_columns(self):
        from openpyxl import load_workbook

        from lavabo.load.senkahomes import COL_AMOUNT, COL_REVIEW, write_orders_workbook
        from lavabo.models import Conversation, ExtractionResult, Source

        conv = Conversation(conversation_id="a", source=Source.ZALO,
                            customer_name="Thảo Nguyễn", messages=[])
        conv.raw = {"order_day": 14, "order_month": 7, "order_number": 1,
                    "extras": [{"kind": "update", "text": "Tổng 12.000",
                                "confidence": "low"}]}
        result = ExtractionResult(conversation_id="a", source=Source.ZALO, values={
            "items": [{"name": "tủ BC52", "quantity": 1}], "address": "TB",
            "total_text": "5.800", "deposit_text": "500k"})

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "out.xlsx"
            write_orders_workbook(out, [conv], {"a": result}, default_year=2026)
            row = list(load_workbook(out).active.iter_rows(min_row=2, values_only=True))[0]

        self.assertEqual(row[6], 5_800_000, "Tổng must be what the order said")
        self.assertEqual(row[8], 500_000, "Cọc must be what the order said")
        self.assertEqual(row[7], 5_300_000, "Xe thu hộ is Tổng - Cọc, unaffected")
        self.assertEqual(row[COL_AMOUNT - 1], "Tổng 12.000.000", "shown, in its own column")
        self.assertEqual(row[COL_REVIEW - 1], "bổ sung — chưa chắc")


if __name__ == "__main__":
    unittest.main(verbosity=2)
