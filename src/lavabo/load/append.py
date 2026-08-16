"""Insert orders into the shop's own workbook, rather than yet another export file.

The separate export stays the default. This is the deliberate second step: open
QUẢN LÝ ĐƠN SENKAHOMES.xlsx, find the month's sheet, and add the orders that are not
already there -- after taking a backup.

Three rules, in order of importance:

1. Back up first, every time. The file is the business's records.
2. Never delete or rewrite an existing row. Rows already in the sheet were put there
   by a person, possibly edited by hand, and this has no way to know which parts were
   deliberate.
3. An order already present is reported, not touched. Matching is by date + customer,
   the only business key the sheet actually carries -- it has no order-number column.

Rule 3 is why this appends rather than updates. Replacing a row group in place means
changing how many rows it occupies when its item count differs, which shifts every row
below it and would silently break anything referring to them. Worth doing later, with
an explicit diff; not worth guessing at now.
"""

from __future__ import annotations

import logging
import shutil
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..models import Conversation, ExtractionResult
from ..money import parse_vnd
from .senkahomes import HEADERS, _address, _items, _order_date, _order_sort_key

log = logging.getLogger(__name__)

MONEY_COLUMNS = {7, 8, 9}
MONEY_FORMAT = "#,##0"
DATE_FORMAT = "DD/MM/YYYY"
THIN = Side(style="thin", color="BFBFBF")
NEW_FILL = PatternFill("solid", fgColor="EAF3EA")     # faint: shows what this run added


def backup(path: Path) -> Path:
    """Timestamped copy beside the original, before anything is written."""
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    target = path.with_name(f"{path.stem}.backup-{stamp}{path.suffix}")
    shutil.copy2(path, target)
    log.info("backup: %s", target.name)
    return target


def sheet_name_for(month: int, year: int) -> str:
    return f"{month:02d}{year}"


def _norm(value: Any) -> str:
    return " ".join(str(value or "").split()).casefold()


def existing_keys(ws) -> set[tuple[str, str]]:
    """(date, customer) for every order group already in the sheet.

    Only rows carrying an STT start a group; the item rows beneath them are blank in
    those columns and must not be read as orders.
    """
    keys: set[tuple[str, str]] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        when, who = row[1], row[2]
        if isinstance(when, datetime):
            when = when.date()
        key = (when.isoformat() if isinstance(when, date) else _norm(when), _norm(who))
        keys.add(key)
    return keys


# Columns that mean "an order lives on this row". Xe thu hộ (H) is excluded on
# purpose: the shop's sheets ship with =G{n}-I{n} pre-filled hundreds of rows down,
# so treating H as content puts the data below the template instead of in it.
CONTENT_COLUMNS = (1, 2, 3, 4, 5, 6, 7, 9, 10, 11, 12)
ORDER_COLUMNS = (1, 2, 3)          # STT / date / customer — only a real order has these


def _has(ws, row: int, columns) -> bool:
    return any(ws.cell(row=row, column=c).value not in (None, "") for c in columns)


def first_free_row(ws) -> int:
    """Where the next order should start.

    The end of the last order group, not the end of the sheet. Sheets carry a
    summary block below the data ("Tổng doanh thu cửa hàng", "=sum(G1:G120)"), and
    appending after *that* would put orders beneath their own totals.
    """
    last_order = 0
    for row in range(2, ws.max_row + 1):
        if _has(ws, row, ORDER_COLUMNS):
            last_order = row
    if not last_order:
        return 2

    # Item rows follow their order with only name/quantity filled.
    row = last_order
    while _has(ws, row + 1, (5, 6)) and not _has(ws, row + 1, ORDER_COLUMNS):
        row += 1
    return row + 1


def collisions(ws, start: int, count: int) -> list[int]:
    """Rows in the write range that already hold something other than the H formula."""
    return [r for r in range(start, start + count) if _has(ws, r, CONTENT_COLUMNS)]


def next_stt(ws, upto: int) -> int:
    highest = 0
    for row in range(2, upto + 1):
        value = ws.cell(row=row, column=1).value
        if isinstance(value, (int, float)):
            highest = max(highest, int(value))
    return highest + 1


def _style(ws, row: int, *, mark_new: bool) -> None:
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        cell.alignment = Alignment(vertical="top", wrap_text=col in (4, 5))
        if col in MONEY_COLUMNS and isinstance(cell.value, (int, float)):
            cell.number_format = MONEY_FORMAT
        if col == 2 and isinstance(cell.value, (date, datetime)):
            cell.number_format = DATE_FORMAT
        if mark_new:
            cell.fill = NEW_FILL


def append_orders(
    workbook: Path,
    conversations: list[Conversation],
    results: dict[str, ExtractionResult],
    *,
    month: int,
    year: int,
    sheet: str | None = None,
    status: str = "New",
    closer: str | None = None,
    dry_run: bool = False,
    mark_new: bool = True,
) -> dict[str, Any]:
    """Add missing orders to the month's sheet. Returns a summary of what happened."""
    if not workbook.exists():
        raise FileNotFoundError(f"{workbook} does not exist")

    name = sheet or sheet_name_for(month, year)
    wb = load_workbook(workbook)

    created_sheet = name not in wb.sheetnames
    ws = wb.create_sheet(name) if created_sheet else wb[name]
    if created_sheet:
        ws.append(HEADERS)
        log.info("created sheet %s", name)
    else:
        header = [ws.cell(row=1, column=i + 1).value for i in range(len(HEADERS))]
        if [_norm(h) for h in header] != [_norm(h) for h in HEADERS]:
            raise ValueError(
                f"sheet {name!r} does not have the expected columns — refusing to write.\n"
                f"  found:    {header}\n  expected: {HEADERS}"
            )

    present = existing_keys(ws)
    orders = sorted(conversations, key=_order_sort_key)

    to_add, already = [], []
    for conv in orders:
        when = _order_date(conv, year)
        key = (when.isoformat() if when else "", _norm(conv.customer_name))
        (already if key in present else to_add).append(conv)

    summary: dict[str, Any] = {
        "sheet": name, "created_sheet": created_sheet,
        "added": len(to_add), "already_present": len(already),
        "rows_written": 0, "backup": None, "start_row": None, "collision": [],
        "already_names": [c.customer_name for c in already],
    }

    start_row = first_free_row(ws)
    planned = sum(
        len(_items((results[c.conversation_id].values
                    if c.conversation_id in results and not results[c.conversation_id].error
                    else {})) or [(None, None)])
        for c in to_add
    )
    summary["start_row"] = start_row
    summary["rows_written"] = planned

    clash = collisions(ws, start_row, planned)
    if clash:
        summary["collision"] = clash[:8]
        summary["rows_written"] = 0
        return summary

    if dry_run or not to_add:
        return summary

    summary["backup"] = str(backup(workbook))
    row = start_row
    stt = next_stt(ws, start_row - 1)

    for conv in to_add:
        res = results.get(conv.conversation_id)
        values = res.values if res and not res.error else {}

        total = parse_vnd(values.get("total_text"))
        deposit = parse_vnd(values.get("deposit_text")) or 0
        collect = (total - deposit) if total is not None else None
        items = _items(values) or [(None, None)]

        first = True
        for item_name, qty in items:
            payload = ([stt, _order_date(conv, year), conv.customer_name, _address(values),
                        item_name, qty, total, collect, deposit, status,
                        values.get("delivery_date_text") or None,
                        conv.raw.get("sender_name") or closer or None]
                       if first else
                       [None, None, None, None, item_name, qty,
                        None, None, None, None, None, None])
            for col, value in enumerate(payload, start=1):
                cell = ws.cell(row=row, column=col)
                # The sheet computes Xe thu hộ itself. Writing a number over
                # =G{n}-I{n} would replace the shop's own formula with a snapshot
                # that stops tracking edits to Tổng or Cọc.
                if col == 8 and isinstance(cell.value, str) and cell.value.startswith("="):
                    continue
                cell.value = value
            _style(ws, row, mark_new=mark_new)
            row += 1
            first = False
        stt += 1

    written = row - start_row
    summary["rows_written"] = written

    wb.save(workbook)
    log.info("appended %d order(s) / %d row(s) to %s", len(to_add), written, name)
    return summary
