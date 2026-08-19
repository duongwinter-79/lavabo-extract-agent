"""Writer for the QUẢN LÝ ĐƠN SENKAHOMES layout.

Differs from the generic writer in shape: an order occupies a GROUP of rows. The
first carries every order-level field, and each further product adds a row where
only the name and quantity are filled -- matching the existing workbook, which
averages 3.23 item rows per order.

Most columns never reach the model. Date, order number and customer come parsed
from the note's header; Xe thu hộ is arithmetic (Tổng - Cọc, verified against 89 of
90 real orders); STT is a counter. Only the address, the line items and the two
money strings are extracted, and the money is converted here rather than by the
model. See docs/06-output-mapping.md.
"""

from __future__ import annotations

import logging
import os
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .. import extras
from ..models import Conversation, ExtractionResult
from ..money import parse_vnd

log = logging.getLogger(__name__)

HEADERS = [
    "STT", "NGÀY CHỐT", "Tên KH", "Địa chỉ", "Tên sản phẩm", "Số lượng",
    "Tổng Tiền hóa đơn", "Xe thu hộ", "Cọc", "Trạng thái", "Ngày hẹn giao",
    "Người chốt đơn",
]

# Export-only. `lavabo append` writes into the shop's own workbook, whose sheets have
# exactly the 12 columns above and formulas positioned around them, so it must keep
# using HEADERS -- widening that list would trip its own column guard. A workbook this
# tool creates from scratch has no such constraint, so revisions get their own columns
# here and nowhere else.
EXTRA_HEADERS = ["Bổ sung", "Số tiền bổ sung", "Cần xem lại"]
EXPORT_HEADERS = HEADERS + EXTRA_HEADERS
COL_EXTRA = len(HEADERS) + 1       # M: the later message, verbatim
COL_AMOUNT = len(HEADERS) + 2      # N: the money it states, parsed but NOT applied
COL_REVIEW = len(HEADERS) + 3      # O: why this row needs a human

MONEY_COLUMNS = {7, 8, 9}          # 1-based: Tổng, Xe thu hộ, Cọc
MONEY_FORMAT = "#,##0"
DATE_FORMAT = "DD/MM/YYYY"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MISSING_FILL = PatternFill("solid", fgColor="FFF2CC")
# Distinct from MISSING_FILL on purpose: pale amber already means "the AI found nothing
# here", and a row needing a human decision is a different problem with a different fix.
REVIEW_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN = Side(style="thin", color="BFBFBF")


# Fields this layout reads out of an extraction. A schema without them yields a workbook
# where only the header-derived columns are filled, which reads as a failed extraction
# rather than the wrong schema being active -- so name the real cause instead.
REQUIRED_FIELDS = ("items", "address", "total_text", "deposit_text")


def missing_schema_fields(schema) -> list[str]:
    have = set(schema.names)
    return [name for name in REQUIRED_FIELDS if name not in have]


def _duplicate_keys(conversations: list[Conversation]) -> set[tuple]:
    """Order keys (day, month, số đơn) held by more than one captured order.

    A safety net independent of the revision sidecar, which only links a version to the
    order it was merged against. Two separate files can still carry the same key -- from
    a capture made before versions were tracked, or from a filename collision -- and
    then nothing connects them: they extract independently, land as two full rows with
    two STTs, and both totals count. That is the one duplicate shape that reaches the
    money silently, so it is detected here from the rows themselves.
    """
    seen: dict[tuple, int] = {}
    for conv in conversations:
        raw = conv.raw or {}
        key = (raw.get("order_day"), raw.get("order_month"), raw.get("order_number"))
        if all(part is not None for part in key):
            seen[key] = seen.get(key, 0) + 1
    return {key for key, count in seen.items() if count > 1}


def _order_sort_key(conv: Conversation) -> tuple:
    raw = conv.raw or {}
    return (raw.get("order_month") or 0, raw.get("order_day") or 0,
            raw.get("order_number") or 0, conv.conversation_id)


def _order_date(conv: Conversation, default_year: int) -> date | None:
    raw = conv.raw or {}
    day, month = raw.get("order_day"), raw.get("order_month")
    if not day or not month:
        return None
    try:
        # Headers almost never carry a year; the capture is month-scoped, so the
        # run's year is the right default rather than something inferred per row.
        return date(raw.get("order_year") or default_year, int(month), int(day))
    except ValueError:
        return None


def _address(values: dict[str, Any]) -> str | None:
    """Address with the phone appended, matching the existing sheet's convention."""
    address = (values.get("address") or "").strip()
    phone = (values.get("phone") or "").strip()
    if address and phone and phone not in address:
        return f"{address} - {phone}"
    return address or phone or None


def _items(values: dict[str, Any]) -> list[tuple[str, Any]]:
    """[(name, quantity)] from the extracted object array, tolerating odd shapes."""
    out: list[tuple[str, Any]] = []
    for entry in values.get("items") or []:
        if isinstance(entry, dict):
            name = (entry.get("name") or "").strip()
            qty = entry.get("quantity")
        else:
            name, qty = str(entry).strip(), None
        if not name:
            continue
        try:
            qty = int(qty) if qty is not None and float(qty) == int(float(qty)) else qty
        except (TypeError, ValueError):
            pass
        out.append((name, qty if qty is not None else 1))
    return out


def write_orders_workbook(
    path: Path,
    conversations: list[Conversation],
    results: dict[str, ExtractionResult],
    *,
    sheet_name: str | None = None,
    default_year: int | None = None,
    default_status: str = "New",
    closer: str | None = None,
) -> Path:
    default_year = default_year or date.today().year
    orders = sorted(conversations, key=_order_sort_key)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name or f"{orders[0].raw.get('order_month', date.today().month):02d}{default_year}" \
        if orders else f"{date.today():%m%Y}"
    ws.append(EXPORT_HEADERS)

    stt = 0
    missing: list[str] = []
    duplicate_keys = _duplicate_keys(orders)

    for conv in orders:
        res = results.get(conv.conversation_id)
        values = res.values if res and not res.error else {}
        if not res or res.error:
            missing.append(conv.conversation_id)

        total = parse_vnd(values.get("total_text"))
        deposit = parse_vnd(values.get("deposit_text")) or 0
        # Not extracted: the existing workbook computes it, and arithmetic here
        # cannot disagree with itself the way a second extraction could.
        collect = (total - deposit) if total is not None else None

        items = _items(values) or [(None, None)]
        stt += 1
        first = True
        updates, versions = extras.summary(conv.raw)
        who = conv.raw.get("sender_name") or closer or None
        is_dupe = (conv.raw.get("order_day"), conv.raw.get("order_month"),
                   conv.raw.get("order_number")) in duplicate_keys
        # "chưa chắc" is a weaker claim than "có bổ sung" and has to read as one: the
        # segmenter is told to keep a revision it cannot classify rather than drop it,
        # and that trade is only sound if the reader can tell a guess from a certainty.
        revision = ("bổ sung — chưa chắc" if extras.uncertain(conv.raw)
                    else "có bổ sung") if updates else ""
        reasons = [r for r in ("trùng số đơn" if is_dupe else "", revision,
                               *(conv.raw.get("flags") or [])) if r]

        for name, qty in items:
            if first:
                ws.append([
                    stt,
                    _order_date(conv, default_year),
                    conv.customer_name,
                    _address(values),
                    name, qty,
                    total, collect, deposit,
                    default_status,
                    values.get("delivery_date_text") or None,
                    # A Zalo OA delivery records who posted the order, so use that in
                    # preference to the run-wide default, which is only a guess.
                    who,
                    # Later messages changing this order, verbatim and uninterpreted,
                    # with any money they state read out beside them so the figure is
                    # visible without opening the cell.
                    "\n\n".join(updates) or None,
                    " · ".join(filter(None, (extras.amounts(u) for u in updates))) or None,
                    ", ".join(reasons) or None,
                ])
                first = False
            else:
                ws.append([None, None, None, None, name, qty,
                           None, None, None, None, None, None, None, None, None])

        # A competing version of the same order gets its own row so both are visible,
        # but carries no money: the shop's sheet totals with =SUMIF and =SUM over these
        # columns, so filling them would count one order twice. STT and Người chốt đơn
        # are left blank for the same reason -- until a human picks a version, this row
        # is evidence, not an order.
        for version in versions:
            ws.append([None,
                       _order_date(conv, default_year),
                       conv.customer_name,
                       None, None, None,
                       None, None, None,
                       None, None, None,
                       version,
                       extras.amounts(version) or None,
                       "2 phiên bản"])

    _finish(ws, len(orders))
    _sheet_run(wb.create_sheet("Run"), orders, results, missing, default_status, closer)

    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    wb.save(tmp)
    os.replace(tmp, path)

    log.info("wrote %s (%d orders, %d rows)", path, len(orders), ws.max_row - 1)
    if missing:
        log.warning("%d order(s) had no successful extraction and are blank beyond the "
                    "header fields", len(missing))
    return path


def _finish(ws, order_count: int) -> None:
    tinted = False
    for i in range(1, len(EXPORT_HEADERS) + 1):
        cell = ws.cell(row=1, column=i)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [6, 13, 20, 42, 46, 9, 17, 14, 12, 13, 14, 15, 52, 26, 14]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for row in ws.iter_rows(min_row=2, max_col=len(EXPORT_HEADERS)):
        for cell in row:
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            cell.alignment = Alignment(
                vertical="top", wrap_text=cell.column in (4, 5, COL_EXTRA))
            if cell.column in MONEY_COLUMNS and isinstance(cell.value, (int, float)):
                cell.number_format = MONEY_FORMAT
            if cell.column == 2 and isinstance(cell.value, date):
                cell.number_format = DATE_FORMAT
        # Flag an order whose total never made it through, so a blank is not read as zero.
        if row[0].value is not None and row[6].value is None:
            row[6].fill = MISSING_FILL
        # Tint the whole order GROUP, not just the row carrying the reason. An order
        # spans one row per line item, and colouring only the first leaves a four-item
        # order three-quarters plain, which reads as a rendering fault rather than a
        # flag. A continuation row has no STT, so the state carries down until the next
        # order starts.
        if row[0].value is not None:
            tinted = bool(row[COL_REVIEW - 1].value)
        if tinted:
            for cell in row:
                cell.fill = REVIEW_FILL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXPORT_HEADERS))}{ws.max_row}"


def _sheet_run(ws, orders, results, missing, status, closer) -> None:
    ok = [r for r in results.values() if not r.error]
    rows = [
        ("generated_at", datetime.now().isoformat(timespec="seconds")),
        ("orders", len(orders)),
        ("extractions_ok", len(ok)),
        ("orders_without_extraction", len(missing)),
        ("default_trạng thái", status),
        ("người chốt đơn", closer or "(not set — see docs/06-output-mapping.md)"),
        ("input_tokens", sum(r.input_tokens for r in results.values())),
        ("output_tokens", sum(r.output_tokens for r in results.values())),
        ("model", next((r.model for r in results.values() if r.model), "-")),
    ]
    ws.append(["key", "value"])
    for key, value in rows:
        ws.append([key, value])
    for i in (1, 2):
        ws.cell(row=1, column=i).fill = HEADER_FILL
        ws.cell(row=1, column=i).font = HEADER_FONT
        ws.column_dimensions[get_column_letter(i)].width = 34
    if missing:
        ws.append([])
        ws.append(["orders without extraction", ""])
        for cid in missing:
            ws.append(["", cid])
