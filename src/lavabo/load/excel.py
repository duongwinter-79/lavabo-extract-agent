"""Excel writer: Data / Sources / Run.

Written to a temp file then atomically replaced, so an Excel window that has the previous
version open never sees a half-written file.
"""

from __future__ import annotations

import logging
import os
from datetime import datetime, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..config import ExtractionSchema
from ..models import Conversation, ExtractionResult

log = logging.getLogger(__name__)

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NULL_FILL = PatternFill("solid", fgColor="FFF2CC")   # amber: model said "not stated"
MAX_WIDTH = 60


def write_workbook(
    path: Path,
    schema: ExtractionSchema,
    conversations: list[Conversation],
    results: dict[str, ExtractionResult],
    *,
    run_meta: dict | None = None,
    display_timezone: str = "UTC",
) -> Path:
    wb = Workbook()
    tz = ZoneInfo(display_timezone)

    _sheet_data(wb.active, schema, conversations, results)
    _sheet_sources(wb.create_sheet("Sources"), conversations, results, tz)
    _sheet_run(wb.create_sheet("Run"), schema, conversations, results, run_meta or {})

    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    wb.save(tmp)
    os.replace(tmp, path)
    log.info("wrote %s (%d rows)", path, len(conversations))
    return path


def _sheet_data(ws, schema: ExtractionSchema, conversations, results) -> None:
    ws.title = "Data"
    headers = ["source", "conversation_id", *schema.names]
    ws.append(headers)

    for conv in conversations:
        res = results.get(conv.conversation_id)
        values = res.values if res else {}
        ws.append([
            conv.source.value,
            conv.conversation_id,
            *(_cell(values.get(name)) for name in schema.names),
        ])

        # Flag cells the model could not fill, so nobody mistakes a blank for a zero.
        row = ws.max_row
        for i, name in enumerate(schema.names, start=3):
            if values.get(name) is None:
                ws.cell(row=row, column=i).fill = NULL_FILL

    _finish(ws, headers)


def _sheet_sources(ws, conversations, results, tz) -> None:
    # order_date / order_no come from the note's header line and are parsed, not
    # inferred, so they are shown here as facts rather than model output.
    headers = ["source", "conversation_id", "customer_name", "order_date", "order_no",
               "customer_handle", "lines", "first_message", "last_message",
               "origin", "extraction_error"]
    # Times shown in local business time, matching the Data sheet and the app UI.
    ws.append(headers)

    for conv in conversations:
        res = results.get(conv.conversation_id)
        ws.append([
            conv.source.value,
            conv.conversation_id,
            conv.customer_name,
            conv.raw.get("order_date_text"),
            conv.raw.get("order_number"),
            conv.customer_handle,
            len(conv.messages),
            _stamp(conv.started_at, tz),
            _stamp(conv.last_message_at, tz),
            conv.origin,
            (res.error if res else "not extracted"),
        ])

    _finish(ws, headers)


def _sheet_run(ws, schema, conversations, results, run_meta: dict) -> None:
    ok = [r for r in results.values() if not r.error]
    filled = sum(
        1 for r in ok for name in schema.names if r.values.get(name) is not None
    )
    total_cells = max(len(ok) * len(schema.names), 1)

    rows = [
        ("generated_at", datetime.now(timezone.utc).isoformat(timespec="seconds")),
        ("schema_version", schema.version),
        ("columns", len(schema.names)),
        ("conversations", len(conversations)),
        ("messages", sum(len(c.messages) for c in conversations)),
        ("extractions_ok", len(ok)),
        ("extractions_failed", len(results) - len(ok)),
        ("fill_rate", f"{filled / total_cells:.1%}"),
        ("input_tokens", sum(r.input_tokens for r in results.values())),
        ("output_tokens", sum(r.output_tokens for r in results.values())),
        ("model", next((r.model for r in results.values() if r.model), "-")),
        *run_meta.items(),
    ]

    ws.append(["key", "value"])
    for key, value in rows:
        ws.append([key, value])
    _finish(ws, ["key", "value"])


def _cell(value):
    """openpyxl accepts only scalars; collapse anything richer to text."""
    if isinstance(value, (list, tuple)):
        return "; ".join(str(v) for v in value)
    if isinstance(value, dict):
        return "; ".join(f"{k}={v}" for k, v in value.items())
    return value


def _stamp(dt, tz=None) -> str:
    if not dt:
        return ""
    return (dt.astimezone(tz) if tz else dt).strftime("%Y-%m-%d %H:%M")


def _finish(ws, headers: list[str]) -> None:
    for i in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=i)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    for i in range(1, len(headers) + 1):
        longest = max(
            (len(str(ws.cell(row=r, column=i).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max(longest + 2, 12), MAX_WIDTH)
